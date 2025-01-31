import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import csv
import datetime
import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import logging
import pandas as pd

# Настройка логирования
logging.basicConfig(filename="app.log", level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

PRODUCTS_FILE = "/storage/emulated/0/Download/Msk Stock Kupi-Flakon/products.csv"
ORDERS_FILE = "/storage/emulated/0/Download/Msk Stock Kupi-Flakon/orders.csv"
STOCKS_FILE = "/storage/emulated/0/Download/Msk Stock Kupi-Flakon/stocks.json"


def load_products(filename=PRODUCTS_FILE):
    try:
        df = pd.read_csv(filename, header=None, names=['name', 'article'], encoding='utf-8', dtype=str)
        return df.to_dict('records')
    except FileNotFoundError:
        open(filename, 'w', encoding='utf-8').close()
        return []
    except Exception as e:
        logging.error(f"Ошибка при загрузке товаров: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при загрузке товаров: {e}")
        return []


def load_stocks(filename=STOCKS_FILE):
    """Загружает остатки товаров из JSON файла."""
    try:
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as file:
                try:
                    return json.load(file)
                except json.JSONDecodeError as e:
                    logging.error(f"Ошибка при загрузке остатков из JSON файла (некорректный JSON): {e}")
                    return {}
        else:
            return {}
    except Exception as e:
        logging.error(f"Ошибка при загрузке остатков: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при загрузке остатков: {e}")
        return {}


def save_stocks(stocks, filename=STOCKS_FILE):
    """Сохраняет остатки товаров в JSON файл."""
    try:
        with open(filename, 'w', encoding='utf-8') as file:
            json.dump(stocks, file, indent=4)
        logging.info(f"Остатки успешно сохранены в {filename}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении остатков: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при сохранении остатков: {e}")

def search_products(products, query):
    """Ищет товары, соответствующие запросу."""
    if not query:
        return products
    query_parts = query.lower().split()
    results = []
    for product in products:
        product_name = product['name'].lower()
        product_article = str(product['article']).lower()
        # Проверяем, содержится ли хотя бы одна часть запроса в названии или артикуле
        if any(part in product_name or part in product_article for part in query_parts):
            results.append(product)
    return results

def display_products(products, listbox, stocks):
    """Отображает товары в Listbox с остатками."""
    listbox.delete(0, tk.END)
    for i, product in enumerate(products):
        stock = stocks.get(product['article'], 0)
        listbox.insert(tk.END, f"{i + 1}. {product['name']} (Арт. {product['article']}, Остаток: {stock})")


def select_product(listbox, products):
    """Позволяет выбрать товар из Listbox."""
    selection = listbox.curselection()
    if selection:
        index = selection[0]
        if 0 <= index < len(products):
            return products[index]
    return None


def get_order_info(window, product, operation_type=""):
    """Запрашивает информацию о заказе через модальное окно."""
    order_window = tk.Toplevel(window)
    order_window.title("Информация о расходах/приходах")

    operation_choices = {
        0: "Расход",
        1: "Приход",
        2: "Инвентаризация"
    }

    if operation_type:
        selected_operation = list(operation_choices.keys())[list(operation_choices.values()).index(operation_type)]
    else:
        selected_operation = 0

    tk.Label(order_window, text="Тип операции:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
    operation_var = tk.IntVar(value=selected_operation)
    for key, value in operation_choices.items():
        tk.Radiobutton(order_window, text=value, variable=operation_var, value=key).grid(row=1, column=key, sticky=tk.W,
                                                                                        padx=5, pady=5)

    order_choices = {
        0: "Ввести вручную",
        1: "Ozon",
        2: "Yandex Market",
        3: "Wildberries",
        4: "DPD",
        5: "CDEK",
        6: "Домодедово",
        7: "Евролог",
        8: "Бийск"
    }

    tk.Label(order_window, text="Источник расхода/прихода:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
    order_source_var = tk.IntVar()
    for key, value in order_choices.items():
        tk.Radiobutton(order_window, text=value, variable=order_source_var, value=key).grid(row=3, column=key,
                                                                                           sticky=tk.W, padx=5, pady=5)

    tk.Label(order_window, text="Номер заказа:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=5)
    order_number_entry = tk.Entry(order_window)
    order_number_entry.grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)

    def on_source_selected():
        if order_source_var.get() == 0:
            order_number_entry.config(state=tk.NORMAL)
        else:
            order_number_entry.config(state=tk.DISABLED)
            order_number_entry.delete(0, tk.END)
            order_number_entry.insert(0, order_choices[order_source_var.get()])

    for key in order_choices:
        tk.Radiobutton(order_window, text=order_choices[key], variable=order_source_var, value=key,
                       command=on_source_selected).grid(row=3, column=key, sticky=tk.W, padx=5, pady=5)

    tk.Label(order_window, text="Дата (ГГГГ-ММ-ДД):").grid(row=5, column=0, sticky=tk.W, padx=5, pady=5)
    date_entry = tk.Entry(order_window)
    date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))  # По умолчанию текущая дата
    date_entry.grid(row=5, column=1, sticky=tk.W, padx=5, pady=5)

    tk.Label(order_window, text="Количество:").grid(row=6, column=0, sticky=tk.W, padx=5, pady=5)
    quantity_entry = tk.Entry(order_window)
    quantity_entry.grid(row=6, column=1, sticky=tk.W, padx=5, pady=5)

    result = {}

    def save_order():
        try:
            operation_type = operation_choices[operation_var.get()]
            order_source_choice = order_source_var.get()
            if order_source_choice == 0:
                order_number = order_number_entry.get()
            else:
                order_number = order_choices[order_source_choice]

            date_str = date_entry.get()
            # Проверяем корректность введённой даты
            try:
                order_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Введите дату в формате ГГГГ-ММ-ДД.")
                return

            quantity = int(quantity_entry.get())
            if quantity < 0:
                messagebox.showerror("Ошибка", "Количество должно быть неотрицательным.")
                return

            result["Operation Type"] = operation_type
            result["Order Number"] = order_number
            result["Date"] = order_date.strftime("%Y-%m-%d")
            result["Quantity"] = quantity

            order_window.destroy()

        except ValueError as e:
            logging.error(f"Ошибка ввода данных заказа: {e}")
            messagebox.showerror("Ошибка", "Неверный формат ввода. Проверьте дату и количество.")

    tk.Button(order_window, text="Сохранить", command=save_order).grid(row=7, column=0, columnspan=2, pady=10)
    order_window.grab_set()
    order_window.wait_window()
    if result:
        return result
    else:
        return None


def write_to_csv(data, filename=ORDERS_FILE):
    """Записывает данные заказа в CSV файл."""
    try:
        with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile,
                                   fieldnames=["Operation Type", "Date", "Order Number", "Product Name", "Product Article",
                                               "Quantity"])
            if csvfile.tell() == 0:
                writer.writeheader()
            for row in data:
                writer.writerow(row)
        messagebox.showinfo("Успех", f"Данные успешно сохранены в {filename}")
        return True
    except Exception as e:
        logging.error(f"Ошибка при записи в файл: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при записи в файл: {e}")
        return False


def update_stocks(order_data, stocks):
    """Обновляет остатки товаров. Обрабатывает ошибки и предупреждения."""
    for order in order_data:
        article = order['Product Article']
        quantity = order['Quantity']
        operation_type = order['Operation Type']

        logging.info(f"Обновление остатков: Артикул={article}, Количество={quantity}, Операция={operation_type}")

        if operation_type == "Приход":
            try:
                stocks[article] = stocks.get(article, 0) + quantity
                logging.info(f"Приход: Остаток для {article} теперь {stocks[article]}")
            except TypeError as e:
                logging.error(f"Ошибка в update_stocks: Некорректное количество: {quantity} для товара {order['Product Name']} (Арт. {article}): {e}")
                messagebox.showerror("Ошибка", f"Некорректное количество: {quantity} для товара {order['Product Name']} (Арт. {article})")
                return False
        elif operation_type == "Расход":
            try:
                stocks[article] = stocks.get(article, 0) - quantity
                logging.info(f"Расход: Остаток для {article} теперь {stocks[article]}")
            except TypeError as e:
                logging.error(f"Ошибка в update_stocks: Некорректное количество: {quantity} для товара {order['Product Name']} (Арт. {article}): {e}")
                messagebox.showerror("Ошибка", f"Некорректное количество: {quantity} для товара {order['Product Name']} (Арт. {article})")
                return False
        elif operation_type == "Инвентаризация":
            try:
                stocks[article] = quantity
                logging.info(f"Инвентаризация: Остаток для {article} теперь {stocks[article]}")
            except TypeError as e:
                logging.error(f"Ошибка в update_stocks: Некорректное количество: {quantity} для товара {order['Product Name']} (Арт. {article}): {e}")
                messagebox.showerror("Ошибка", f"Некорректное количество: {quantity} для товара {order['Product Name']} (Арт. {article})")
                return False
        else:
            messagebox.showwarning("Ошибка", f"Неизвестный тип операции: {operation_type} для товара {order['Product Name']} (Арт. {article})")
            logging.warning(f"Неизвестный тип операции: {operation_type} для товара {order['Product Name']} (Арт. {article})")
    return True

def generate_report(start_date, end_date, selected_product, query, products, filename=ORDERS_FILE):
    """Генерирует общий отчет по операциям для выбранного товара и периода."""
    report_data = []
    try:
        with open(filename, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                try:
                    order_date = datetime.datetime.strptime(row['Date'], "%Y-%m-%d").date()
                except (ValueError, TypeError) as e:
                    logging.error(f"Некорректный формат даты: {row['Date']} в файле {filename}, {e}")
                    continue  # Пропускаем строку с некорректной датой
                if start_date <= order_date <= end_date:
                    if selected_product:
                        # Фильтруем по выбранному товару
                        if row["Product Article"] == selected_product["article"] and row["Product Name"] == selected_product["name"]:
                            report_data.append(row)
                    elif query:
                        # Фильтруем по запросу, если выбранный товар не указан
                        filtered_products = search_products(products, query)
                        if row["Product Article"] in [product["article"] for product in filtered_products]:
                            report_data.append(row)
                    else:
                        # Если нет выбранного товара и запроса, включаем все товары
                        report_data.append(row)
    except FileNotFoundError:
        logging.error(f"Файл {filename} не найден при генерации отчёта")
        messagebox.showerror("Ошибка", f"Файл {filename} не найден.")
        return None
    except Exception as e:
        logging.error(f"Ошибка при генерации отчета: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка при генерации отчета: {e}")
        return None
    return report_data


def display_report(report_data, products, initial_query=""):
    """Отображает отчет в новом окне."""
    if not report_data:
        messagebox.showinfo("Отчет", "Нет данных за выбранный период.")
        return []

    report_window = tk.Toplevel()
    report_window.title("Общий отчет по операциям")
    report_window.geometry("1200x700")

    # --- Filters Frame ---
    filter_frame = tk.Frame(report_window)
    filter_frame.pack(fill=tk.X, padx=5, pady=5)

    # --- Operation Type Filter
    tk.Label(filter_frame, text="Фильтр по типу операции:").pack(side=tk.LEFT)
    operation_types = ["Приход", "Расход", "Инвентаризация"]
    operation_var = tk.StringVar(value="Все")
    operation_var.set("Все")
    operation_menu = tk.OptionMenu(filter_frame, operation_var, "Все", *operation_types)
    operation_menu.pack(side=tk.LEFT, padx=5)

    # --- Search Frame ---
    search_frame = tk.Frame(report_window)
    search_frame.pack(fill=tk.X, padx=5, pady=5)

    tk.Label(search_frame, text="Поиск в отчете:").pack(side=tk.LEFT)
    search_entry = tk.Entry(search_frame)
    search_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
    search_entry.insert(0, initial_query)

    search_field_label = tk.Label(search_frame, text="Поле поиска:")
    search_field_label.pack(side=tk.LEFT, padx=5)
    search_fields = ["Название", "Артикул", "Номер заказа", "Тип операции"]
    search_field_var = tk.StringVar(value="Название")
    search_field_menu = tk.OptionMenu(search_frame, search_field_var, *search_fields)
    search_field_menu.pack(side=tk.LEFT, padx=5)

    # --- Treeview (Table)
    columns = ["Дата", "Номер заказа", "Тип операции", "Название товара", "Артикул товара", "Количество", "Итог"]
    tree = ttk.Treeview(report_window, columns=columns, show="headings")
    for col in columns:
        tree.heading(col, text=col, command=lambda c=col: sort_report(c))

    # Настройка ширины колонок
    tree.column("Дата", width=100)  # Ширина колонки "Дата"
    tree.column("Номер заказа", width=100)  # Ширина колонки "Номер заказа"
    tree.column("Тип операции", width=100)  # Ширина колонки "Тип операции"
    tree.column("Название товара", width=300)  # Увеличиваем ширину колонки "Название товара"
    tree.column("Артикул товара", width=100)  # Ширина колонки "Артикул товара"
    tree.column("Количество", width=100)  # Ширина колонки "Количество"
    tree.column("Итог", width=100)  # Ширина колонки "Итог"

    tree.pack(expand=True, fill=tk.BOTH)

    # --- Кнопка сохранения в Excel
    def on_save_report():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            # Передаем отфильтрованные данные в функцию сохранения
            save_report_to_xlsx(filtered_data, file_path, products)

    def on_save_report_txt():
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        if file_path:
            # Передаем отфильтрованные данные в функцию сохранения
            save_report_to_txt(filtered_data, file_path)

    save_excel_button = tk.Button(report_window, text="Сохранить отчёт в Excel", command=on_save_report)
    save_excel_button.pack(pady=5)

    save_txt_button = tk.Button(report_window, text="Сохранить отчёт в TXT", command=on_save_report_txt)
    save_txt_button.pack(pady=5)

    current_sort_column = None
    sort_direction = False  # True - ascending, False - descending
    filtered_data = report_data  # Изначально отфильтрованные данные равны исходным

    def sort_report(column):
        nonlocal current_sort_column, sort_direction
        if current_sort_column == column:
            sort_direction = not sort_direction
        else:
            current_sort_column = column
            sort_direction = False
        update_report_display(search_entry.get(), operation_var.get())

    def update_report_display(query, operation_filter):
        nonlocal filtered_data

        # Очищаем таблицу
        for item in tree.get_children():
            tree.delete(item)

        # Фильтрация данных по типу операции
        if operation_filter != "Все":
            filtered_data = [row for row in report_data if row["Operation Type"] == operation_filter]
        else:
            filtered_data = report_data

        # Фильтрация данных по запросу
        if query:
            query_field = search_field_var.get()
            query_words = set(query.lower().split())
            temp_filtered_data = []
            for row in filtered_data:
                if query_field == "Название":
                    row_str = str(row["Product Name"]).lower()
                elif query_field == "Артикул":
                    row_str = str(row["Product Article"]).lower()
                elif query_field == "Номер заказа":
                    row_str = str(row["Order Number"]).lower()
                elif query_field == "Тип операции":
                    row_str = str(row["Operation Type"]).lower()
                else:
                    row_str = ' '.join(str(value).lower() for value in row.values())

                row_words = set(row_str.split())
                if query_words.issubset(row_words):
                    temp_filtered_data.append(row)
            filtered_data = temp_filtered_data

        # Сортировка данных по дате
        if current_sort_column:
            def get_sort_key(row):
                if current_sort_column == "Дата":
                    return datetime.datetime.strptime(row['Date'], "%Y-%m-%d").date()
                return row.get(current_sort_column, "")
            filtered_data.sort(key=get_sort_key, reverse=sort_direction)

        # Группировка данных по артикулу товара
        grouped_data = {}
        for row in filtered_data:
            article = row["Product Article"]
            if article not in grouped_data:
                grouped_data[article] = []
            grouped_data[article].append(row)

        # Отображение данных в таблице
        for article, items in grouped_data.items():
            total_in = 0  # Сумма приходов
            total_out = 0  # Сумма расходов
            inventory_qty = None  # Значение инвентаризации
            product_name = ""  # Название товара

            # Сортируем операции по дате
            items.sort(key=lambda x: datetime.datetime.strptime(x['Date'], "%Y-%m-%d").date())

            # Переменная для хранения текущего остатка
            current_stock = 0

            for item in items:
                product_name = item["Product Name"]
                quantity = int(item["Quantity"])

                # Если это инвентаризация, устанавливаем текущий остаток
                if item["Operation Type"] == "Инвентаризация":
                    inventory_qty = quantity
                    current_stock = inventory_qty
                elif item["Operation Type"] == "Приход":
                    current_stock += quantity
                    total_in += quantity
                elif item["Operation Type"] == "Расход":
                    current_stock -= quantity
                    total_out += quantity

                # Добавляем строку с операцией
                tree.insert("", tk.END, values=[
                    item['Date'],
                    item['Order Number'],
                    item['Operation Type'],
                    item['Product Name'],
                    item['Product Article'],
                    item['Quantity'],
                    current_stock  # Текущий остаток после операции
                ])

            # Рассчитываем итоговый остаток
            if inventory_qty is not None:
                # Если была инвентаризация, итоговый остаток равен текущему остатку
                final_stock = current_stock
            else:
                # Если инвентаризации не было, остаток = приходы - расходы
                final_stock = total_in - total_out

            # Отладочный код: выводим итоговый остаток
            print(f"\nТовар: {product_name}, Артикул: {article}")
            print(f"Приход: {total_in}, Расход: {total_out}, Инвентаризация: {inventory_qty}")
            print(f"Итоговый остаток: {final_stock}")

            # Добавляем строку с итогом для каждого товара
            tree.insert("", tk.END, values=[
                "", "", "",
                f"Итог для {product_name} (Арт. {article})",
                "",
                "",
                final_stock  # Итоговый остаток
            ])

            # Добавляем горизонтальную линию после каждого товара
            tree.insert("", tk.END, values=["-" * 30, "-" * 30, "-" * 30, "-" * 100, "-" * 30, "-" * 30, "-" * 30])

        # Отладочный код: выводим итоговые данные в таблице
        print("\nИтоговые данные в таблице:")
        for item in tree.get_children():
            print(tree.item(item, "values"))

        print("--- Конец отладки ---\n")

    def on_search():
        query = search_entry.get()
        operation_filter = operation_var.get()
        update_report_display(query, operation_filter)

    def on_reset_search():
        search_entry.delete(0, tk.END)
        operation_var.set("Все")
        update_report_display("", "Все")

    tk.Button(search_frame, text="Найти", command=on_search).pack(side=tk.LEFT, padx=5)
    tk.Button(search_frame, text="Сбросить поиск", command=on_reset_search).pack(side=tk.LEFT, padx=5)

    update_report_display(initial_query, "Все")  # Первичное отображение отчёта
    return report_data

def save_report_to_txt(report_data, filename):
    """Сохраняет отчет в TXT файл."""
    try:
        with open(filename, 'w', encoding='utf-8') as file:
            header = ["Дата", "Номер заказа", "Тип операции", "Название товара", "Артикул товара", "Количество"]
            file.write(f"{' | '.join(header)}\n")
            file.write("-" * 180 + "\n")

            grouped_data = {}
            for row in report_data:
                article = row["Product Article"]
                if article not in grouped_data:
                    grouped_data[article] = []
                grouped_data[article].append(row)

            for article, items in grouped_data.items():
                total_in = 0
                total_out = 0
                product_name = ""
                inventory_qty = None
                has_inventory = False

                for item in items:
                    product_name = item["Product Name"]
                    quantity = int(item["Quantity"])
                    file.write(
                        f"{item['Date']} | {item['Order Number']} | {item['Operation Type']} | {item['Product Name']} | {item['Product Article']} | {item['Quantity']}\n")
                    if item["Operation Type"] == "Приход":
                        total_in += quantity
                    elif item["Operation Type"] == "Расход":
                        total_out += quantity
                    elif item["Operation Type"] == "Инвентаризация":
                        inventory_qty = quantity
                        has_inventory = True
                file.write("-" * 180 + "\n")
                if has_inventory:
                    file.write(
                        f"Итого для {product_name} (Арт. {article}): Приход: {total_in}, Расход: {total_out}, Инвентаризация: {inventory_qty}, Итог: {inventory_qty}\n")
                else:
                    difference = total_in - total_out
                    file.write(
                        f"Итого для {product_name} (Арт. {article}): Приход: {total_in}, Расход: {total_out}, Итог: {difference}\n")
                file.write("-" * 180 + "\n")

        messagebox.showinfo("Успех", f"Отчет сохранен в {filename}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении отчета в TXT: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка при сохранении отчета в TXT: {e}")


def get_report_params(products, orders_file=ORDERS_FILE):
    """Запрашивает у пользователя параметры для отчета."""
    report_window = tk.Toplevel()
    report_window.title("Параметры отчета")
    report_window.geometry("1200x800")

    tk.Label(report_window, text="Начальная дата (ГГГГ-ММ-ДД):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
    start_date_entry = tk.Entry(report_window)
    start_date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))  # По умолчанию текущая дата
    start_date_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

    tk.Label(report_window, text="Конечная дата (ГГГГ-ММ-ДД):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
    end_date_entry = tk.Entry(report_window)
    end_date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))  # По умолчанию текущая дата
    end_date_entry.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)

    tk.Label(report_window, text="Выбрать товар (оставьте пустым для всех):").grid(row=2, column=0, sticky=tk.W, padx=5,
                                                                                  pady=5)
    product_listbox = tk.Listbox(report_window, width=100, height=30)
    product_listbox.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

    # 1. Загружаем остатки
    stocks = load_stocks()

    result = {}
    # 2. Передаем остатки в display_products
    display_products(products, product_listbox, stocks)  # Вот здесь добавляем stocks

    def on_product_select(event):
        selected_product = select_product(product_listbox, products)
        if selected_product:
            last_inventory_date = get_last_inventory_date(selected_product, orders_file)
            if last_inventory_date:
                start_date_entry.delete(0, tk.END)
                start_date_entry.insert(0, last_inventory_date.strftime("%Y-%m-%d"))
            else:
                start_date_entry.delete(0, tk.END)
                start_date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))

    product_listbox.bind("<<ListboxSelect>>", on_product_select)

    def on_submit():
        start_date_str = start_date_entry.get()
        end_date_str = end_date_entry.get()
        selected_product = select_product(product_listbox, products)

        try:
            # Проверяем корректность введённых дат
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()

            if start_date > end_date:
                messagebox.showerror("Ошибка", "Начальная дата не может быть позже конечной.")
                return

            result["start_date"] = start_date
            result["end_date"] = end_date
            if selected_product:
                result["selected_product"] = selected_product
            else:
                result["selected_product"] = None

            report_window.destroy()

        except ValueError as e:
            logging.error(f"Неверный формат даты отчета: {e}")
            messagebox.showerror("Ошибка", "Неверный формат даты. Введите в формате ГГГГ-ММ-ДД.")

    tk.Button(report_window, text="Сформировать отчет", command=on_submit).grid(row=3, column=0, columnspan=2, pady=10)
    report_window.grab_set()
    report_window.wait_window()
    if result:
        return result
    else:
        return None


def get_last_inventory_date(product, filename=ORDERS_FILE):
    """Получает дату последней инвентаризации для товара."""
    last_inventory_date = None
    try:
        with open(filename, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                if row["Product Article"] == product["article"] and row["Product Name"] == product["name"] and row[
                    "Operation Type"] == "Инвентаризация":
                    try:
                        order_date = datetime.datetime.strptime(row['Date'], "%Y-%m-%d").date()
                        if last_inventory_date is None or order_date > last_inventory_date:
                            last_inventory_date = order_date
                    except (ValueError, TypeError) as e:
                        logging.error(
                            f"Ошибка даты в файле {filename} при определении даты последней инвентаризации для товара {product['name']}: {e}")
    except FileNotFoundError:
        logging.error(f"Файл {filename} не найден при поиске даты последней инвентаризации.")
        messagebox.showerror("Ошибка", f"Файл {filename} не найден.")
    except Exception as e:
        logging.error(f"Ошибка при поиске даты последней инвентаризации: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка при поиске даты последней инвентаризации: {e}")
    return last_inventory_date


def save_report_to_xlsx(report_data, filename, products):
    """Сохраняет отчет в Excel файл, включая итоговую разницу на отдельном листе."""
    try:
        workbook = openpyxl.Workbook()
        report_sheet = workbook.active
        report_sheet.title = "Отчет"

        # Записываем данные отчета
        header = ["Дата", "Номер заказа", "Тип операции", "Название товара", "Артикул товара", "Количество"]
        report_sheet.append(header)
        for row in report_data:
            report_sheet.append(
                [row['Date'], row['Order Number'], row['Operation Type'], row['Product Name'], row['Product Article'],
                 row['Quantity']])

        summary_sheet = workbook.create_sheet("Итог")
        summary_sheet.title = "Итог"

        summary_header = ["Артикул", "Название", "Приход", "Расход", "Инвентаризация", "Итог"]
        summary_sheet.append(summary_header)

        grouped_data = {}
        for row in report_data:
            article = row["Product Article"]
            if article not in grouped_data:
                grouped_data[article] = []
            grouped_data[article].append(row)

        for article, items in grouped_data.items():
            total_in = 0
            total_out = 0
            product_name = ""
            inventory_qty = None
            has_inventory = False

            for item in items:
                product_name = item["Product Name"]
                quantity = int(item["Quantity"])
                if item["Operation Type"] == "Приход":
                    total_in += quantity
                elif item["Operation Type"] == "Расход":
                    total_out += quantity
                elif item["Operation Type"] == "Инвентаризация":
                    inventory_qty = quantity
                    has_inventory = True

            if has_inventory:
                difference = inventory_qty
            else:
                difference = total_in - total_out

            summary_data = [article, product_name, total_in, total_out, inventory_qty if inventory_qty is not None else "",
                            difference]
            summary_sheet.append(summary_data)

        workbook.save(filename)
        messagebox.showinfo("Успех", f"Отчет сохранен в {filename}")

    except Exception as e:
        logging.error(f"Ошибка при сохранении отчета в excel: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка при сохранении отчета в excel: {e}")


def generate_stock_report(products, stocks):
    """Генерирует отчет по остаткам товаров."""
    stock_report_data = []
    for product in products:
        article = product['article']
        name = product['name']
        stock = stocks.get(article, 0)
        stock_report_data.append({"Название товара": name, "Артикул товара": article, "Остаток": stock})
    return stock_report_data


def display_stock_report(stock_report_data):
    """Отображает отчет по остаткам в новом окне."""
    if not stock_report_data:
        messagebox.showinfo("Отчет", "Нет данных по остаткам.")
        return

    stock_report_window = tk.Toplevel()
    stock_report_window.title("Отчет по остаткам товаров")
    stock_report_window.geometry("800x600")

    text_area = tk.Text(stock_report_window, wrap=tk.WORD)
    text_area.pack(expand=True, fill=tk.BOTH)

    header = ["Название товара", "Артикул товара", "Остаток"]
    text_area.insert(tk.END, f"{' | '.join(header)}\n")
    text_area.insert(tk.END, "-" * 60 + "\n")

    for item in stock_report_data:
        text_area.insert(tk.END, f"{item['Название товара']} | {item['Артикул товара']} | {item['Остаток']}\n")

    text_area.config(state=tk.DISABLED)

    def on_save_stock_report():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            save_stock_report_to_xlsx(stock_report_data, file_path)

    save_button = tk.Button(stock_report_window, text="Сохранить остатки в Excel", command=on_save_stock_report)
    save_button.pack(pady=10)


def save_stock_report_to_xlsx(stock_report_data, filename):
    """Сохраняет отчет по остаткам в Excel файл."""
    try:
        workbook = openpyxl.Workbook()
        stock_sheet = workbook.active
        stock_sheet.title = "Остатки"

        header = ["Название товара", "Артикул товара", "Остаток"]
        stock_sheet.append(header)

        for item in stock_report_data:
            stock_sheet.append([item['Название товара'], item['Артикул товара'], item['Остаток']])

        workbook.save(filename)
        messagebox.showinfo("Успех", f"Отчет по остаткам сохранен в {filename}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении отчета по остаткам в excel: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка при сохранении отчета по остаткам в excel: {e}")


def load_new_stocks(window, stocks, products):
    """Позволяет загрузить новые остатки из файла Excel."""
    file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Берем активный лист

        header = [cell.value for cell in sheet[1]]  # Читаем заголовки из 1й строки

        if not header or "Артикул товара" not in header or "Остаток" not in header or "Название товара" not in header:
            messagebox.showerror("Ошибка",
                                "Неверный формат файла. Проверьте наличие столбцов 'Название товара', 'Артикул товара' и 'Остаток'.")
            return

        name_col = header.index("Название товара")
        article_col = header.index("Артикул товара")
        stock_col = header.index("Остаток")

        new_stocks = {}
        new_products = []  # Список для хранения новых товаров
        order_data = []  # Список для хранения данных о приходе/инвентаризации

        for row_num in range(2, sheet.max_row + 1):
            row = [cell.value for cell in sheet[row_num]]
            if len(row) > max(name_col, article_col, stock_col):
                name = str(row[name_col]).strip()
                article = str(row[article_col]).strip()
                try:
                    stock = int(row[stock_col])

                    # Проверяем, есть ли товар в списке products
                    product_exists = False
                    for p in products:
                        if p['name'] == name and p['article'] == article:
                            product_exists = True
                            break

                    if product_exists:
                        new_stocks[article] = stock
                    else:
                        # Добавляем новый товар в список products
                        new_product = {"name": name, "article": article}
                        products.append(new_product)
                        new_products.append(new_product)
                        new_stocks[article] = stock

                    # Создаем запись о приходе или инвентаризации
                    order_info = {
                        "Operation Type": "Инвентаризация",  # Можно изменить на "Приход", если это приход
                        "Order Number": "Загрузка из файла",
                        "Date": datetime.date.today().strftime("%Y-%m-%d"),
                        "Product Name": name,
                        "Product Article": article,
                        "Quantity": stock
                    }
                    order_data.append(order_info)

                except ValueError as e:
                    logging.error(f"Неверный формат данных в файле Excel в строке {row_num} загрузки остатков: {e}")
                    messagebox.showerror("Ошибка", f"Неверный формат данных в строке {row_num}. Проверьте правильность ввода остатка.")
                    return

        # Если есть новые товары, сохраняем их в products.csv
        if new_products:
            try:
                with open(PRODUCTS_FILE, 'a', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    for product in new_products:
                        writer.writerow([product['name'], product['article']])
                messagebox.showinfo("Успех", f"Добавлено {len(new_products)} новых товаров в список.")
            except Exception as e:
                logging.error(f"Ошибка при сохранении новых товаров: {e}")
                messagebox.showerror("Ошибка", f"Ошибка при сохранении новых товаров: {e}")

        # Обновляем остатки и сохраняем
        stocks.update(new_stocks)
        save_stocks(stocks)

        # Записываем данные о приходе/инвентаризации в orders.csv
        if order_data:
            write_to_csv(order_data)

        messagebox.showinfo("Успех", "Остатки успешно загружены из файла.")

    except FileNotFoundError:
        logging.error(f"Файл не найден при загрузке остатков: {file_path}")
        messagebox.showerror("Ошибка", f"Файл {file_path} не найден.")
    except Exception as e:
        logging.error(f"Произошла ошибка при загрузке остатков: {e}")
        messagebox.showerror("Ошибка", f"Произошла ошибка при загрузке файла. {e}")

def add_product(products, product_name, product_article):
    """Добавляет новый товар в список."""
    try:
        # Проверяем, существует ли уже товар с таким артикулом
        for product in products:
            if product['article'] == product_article:
                messagebox.showwarning("Предупреждение", f"Товар с артикулом {product_article} уже существует.")
                return False
        
        # Добавляем новый товар
        new_product = {"name": product_name, "article": product_article}
        products.append(new_product)
        
        # Сохраняем изменения в файл
        with open(PRODUCTS_FILE, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([product_name, product_article])
        
        messagebox.showinfo("Успех", f"Товар '{product_name}' (артикул: {product_article}) успешно добавлен.")
        return True
    except Exception as e:
        logging.error(f"Ошибка при добавлении товара: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при добавлении товара: {e}")
        return False

def delete_product(products, product_article):
    """Удаляет товар из списка по артикулу."""
    try:
        # Ищем товар по артикулу
        product_to_delete = None
        for product in products:
            if product['article'] == product_article:
                product_to_delete = product
                break
        
        if product_to_delete:
            # Удаляем товар из списка
            products.remove(product_to_delete)
            
            # Перезаписываем файл с обновленным списком товаров
            with open(PRODUCTS_FILE, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                for product in products:
                    writer.writerow([product['name'], product['article']])
            
            messagebox.showinfo("Успех", f"Товар с артикулом {product_article} успешно удален.")
            return True
        else:
            messagebox.showwarning("Предупреждение", f"Товар с артикулом {product_article} не найден.")
            return False
    except Exception as e:
        logging.error(f"Ошибка при удалении товара: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при удалении товара: {e}")
        return False

def on_add_product(window, products, product_listbox, stocks):
    """Обрабатывает добавление нового товара."""
    product_name = simpledialog.askstring("Добавить товар", "Введите название товара:")
    if not product_name:
        return
    
    product_article = simpledialog.askstring("Добавить товар", "Введите артикул товара:")
    if not product_article:
        return
    
    if add_product(products, product_name, product_article):
        display_products(products, product_listbox, stocks)

def on_delete_product(window, products, product_listbox, stocks):
    """Обрабатывает удаление товара."""
    product_article = simpledialog.askstring("Удалить товар", "Введите артикул товара для удаления:")
    if not product_article:
        return
    
    if delete_product(products, product_article):
        display_products(products, product_listbox, stocks)

def edit_product(products, product_article, new_name=None, new_article=None):
    """Изменяет название и/или артикул товара."""
    try:
        # Ищем товар по артикулу
        product_to_edit = None
        for product in products:
            if product['article'] == product_article:
                product_to_edit = product
                break
        
        if product_to_edit:
            # Если передано новое название, обновляем его
            if new_name:
                product_to_edit['name'] = new_name
            
            # Если передан новый артикул, обновляем его
            if new_article:
                # Проверяем, не существует ли уже товар с таким артикулом
                for product in products:
                    if product['article'] == new_article:
                        messagebox.showwarning("Предупреждение", f"Товар с артикулом {new_article} уже существует.")
                        return False
                product_to_edit['article'] = new_article
            
            # Перезаписываем файл с обновленным списком товаров
            with open(PRODUCTS_FILE, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                for product in products:
                    writer.writerow([product['name'], product['article']])
            
            messagebox.showinfo("Успех", f"Товар с артикулом {product_article} успешно изменен.")
            return True
        else:
            messagebox.showwarning("Предупреждение", f"Товар с артикулом {product_article} не найден.")
            return False
    except Exception as e:
        logging.error(f"Ошибка при изменении товара: {e}")
        messagebox.showerror("Ошибка", f"Ошибка при изменении товара: {e}")
        return False
        
def on_edit_product(window, products, product_listbox, stocks):
    """Обрабатывает изменение товара."""
    product_article = simpledialog.askstring("Изменить товар", "Введите артикул товара для изменения:")
    if not product_article:
        return
    
    new_name = simpledialog.askstring("Изменить товар", "Введите новое название товара (оставьте пустым, чтобы не изменять):")
    new_article = simpledialog.askstring("Изменить товар", "Введите новый артикул товара (оставьте пустым, чтобы не изменять):")
    
    if edit_product(products, product_article, new_name, new_article):
        display_products(products, product_listbox, stocks)


def main():
    """Основная функция для создания GUI."""
    window = tk.Tk()
    window.title("Учет товаров Склад МСК Купи-Флакон")
    window.geometry("1800x1000")
    #window.iconbitmap("icon.ico")

    products = load_products()
    stocks = load_stocks()
    order_data = []
    report_data = []
    stock_report_data = []

    tk.Label(window, text="Поиск товара:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
    search_entry = tk.Entry(window, width=50, font=("Arial", 14))  # Увеличиваем ширину и шрифт
    search_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)

    product_listbox = tk.Listbox(window, width=180, height=30)
    product_listbox.grid(row=1, column=0, columnspan=2, padx=5, pady=5)

    display_products(products, product_listbox, stocks)

    def on_search():
        query = search_entry.get()
        search_results = search_products(products, query)
        display_products(search_results, product_listbox, stocks)

    def on_reset_search():
        search_entry.delete(0, tk.END)
        display_products(products, product_listbox, stocks)

    def on_select_and_order():
        query = search_entry.get()
        search_results = search_products(products, query)
        selected_product = select_product(product_listbox, search_results)
        if selected_product:
            order_info = get_order_info(window, selected_product)
            if order_info:
                order_info["Product Name"] = selected_product['name']
                order_info["Product Article"] = selected_product['article']
                order_data.append(order_info)
        else:
            messagebox.showinfo("Инфо", "Товар не выбран.")

    def on_inventory():
        query = search_entry.get()
        search_results = search_products(products, query)
        selected_product = select_product(product_listbox, search_results)
        if selected_product:
            order_info = get_order_info(window, selected_product, operation_type="Инвентаризация")
            if order_info:
                order_info["Product Name"] = selected_product['name']
                order_info["Product Article"] = selected_product['article']
                order_data.append(order_info)
        else:
            messagebox.showinfo("Инфо", "Товар не выбран.")

    tk.Button(window, text="Найти", command=on_search).grid(row=0, column=2, padx=5, pady=5)
    tk.Button(window, text="Сбросить поиск", command=on_reset_search).grid(row=0, column=3, padx=5, pady=5)
    tk.Button(window, text="Выбрать и добавить в расход/приход", command=on_select_and_order).grid(row=2, column=0,
                                                                                                  columnspan=2, pady=10)
    tk.Button(window, text="Инвентаризация", command=on_inventory).grid(row=3, column=0, columnspan=2, pady=10)

        # Добавляем кнопки для добавления и удаления товаров
    tk.Button(window, text="Добавить товар", command=lambda: on_add_product(window, products, product_listbox, stocks)).grid(row=2, column=2, columnspan=2, pady=10)
    tk.Button(window, text="Удалить товар", command=lambda: on_delete_product(window, products, product_listbox, stocks)).grid(row=3, column=2, columnspan=2, pady=10)
    tk.Button(window, text="Изменить товар", command=lambda: on_edit_product(window, products, product_listbox, stocks)).grid(row=4, column=2, columnspan=2, pady=10)
    
    def on_save():
        nonlocal stocks
        if not order_data:
            messagebox.showinfo("Инфо", "Нет данных для сохранения.")
            return

        if update_stocks(order_data, stocks):
            save_stocks(stocks)
            if write_to_csv(order_data):
                order_data.clear()
                display_products(search_products(products, search_entry.get()), product_listbox, stocks)
        else:
            messagebox.showinfo("Ошибка", "Ошибка при обновлении остатков.")

    def on_show_report():
        nonlocal report_data
        report_params = get_report_params(products)
        if report_params:
            report_data = generate_report(report_params["start_date"], report_params["end_date"],
                                         report_params["selected_product"], "", products)
            if report_data:
                display_report(report_data, products)
            else:
                messagebox.showinfo("Отчет", "Нет данных за выбранный период.")

    def on_show_stock_report():
        nonlocal stock_report_data
        stock_report_data = generate_stock_report(products, stocks)
        display_stock_report(stock_report_data)

    def on_load_new_stocks():
        nonlocal stocks
        load_new_stocks(window, stocks, products)
        display_products(search_products(products, search_entry.get()), product_listbox, stocks)

    tk.Button(window, text="Сохранить расход/приход", command=on_save).grid(row=4, column=0, columnspan=2, pady=10)
    tk.Button(window, text="Общий отчет", command=on_show_report).grid(row=2, column=1, columnspan=2, pady=10)
    tk.Button(window, text="Отчет по остаткам", command=on_show_stock_report).grid(row=3, column=1, columnspan=2, pady=10)
    tk.Button(window, text="Загрузить новые остатки", command=on_load_new_stocks).grid(row=4, column=1, columnspan=2, pady=10)

    display_products(products, product_listbox, stocks)  # Отображаем все товары при запуске
    window.mainloop()


if __name__ == "__main__":
    main()
